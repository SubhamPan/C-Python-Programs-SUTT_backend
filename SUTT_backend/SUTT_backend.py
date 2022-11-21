import pandas as pd
import openpyxl


workbook = openpyxl.load_workbook("Names_ID.xlsx")
sheet = workbook.active
first_row = ["BITS ID", "Name", "BITS email", "Branch"] # The row where we stock the name of the column



# for col in range(1, sheet.max_column+1):
#     first_row.append(sheet.cell(row=1, column=col).value)



data =[]


for row in range(2, sheet.max_row+1):
    dictionary = {}

    #BITS ID and Name from excel
    for col in range(1, 3):
        dictionary[first_row[col-1]]=sheet.cell(row=row,column=col).value
    


    #BITS email 
    dictionary[first_row[2]] = "f" + str(sheet.cell(row = row, column = 1).value)[0:4] + str(sheet.cell(row = row, column = 1).value)[8:12] + "@pilani.bits-pilani.ac.in"

    

    #branch (checking for branch code from bits ID iven in excel)
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "AA"):
        dictionary[first_row[3]] = "ECE"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "AB"):
        dictionary[first_row[3]] = "Manufacturing"

    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A1"):
        dictionary[first_row[3]] = "Chemical"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A2"):
        dictionary[first_row[3]] = "Civil"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A3"):
        dictionary[first_row[3]] = "EEE"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A4"):
        dictionary[first_row[3]] = "Mechanical"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A5"):
        dictionary[first_row[3]] = "B Pharm"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A7"):
        dictionary[first_row[3]] = "CSE"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "A8"):
        dictionary[first_row[3]] = "ENI"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "B1"):
        dictionary[first_row[3]] = "Msc Bio"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "B2"):
        dictionary[first_row[3]] = "Msc Chem"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "B3"):
        dictionary[first_row[3]] = "Msc Eco"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "B4"):
        dictionary[first_row[3]] = "Msc Mathematics"
    
    if(str(sheet.cell(row = row, column = 1).value)[4:6] == "B5"):
        dictionary[first_row[3]] = "Msc Physics"
   
    
    
    #update the dictionary to the data
    data.append(dictionary)

    
print (data)



# dataframe1 = pd.read_excel('Names_ID.xlsx')
# print(dataframe1)